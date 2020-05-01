import json
import os
from header_node import HeaderNode
from jsonschema import Draft7Validator, RefResolver
from loaded_node_object_info import LoadedNodeObjectInfo
from openpyxl import load_workbook
from type_node import TypeNode, TypeNodeParseType

class DataExporterConfig:
    def __init__(self, config_json):
        self.schema_dir_path = config_json['schema_dir_path']
        self.table_dir_path = config_json['table_dir_path']
        self.asset_dir_path = config_json['asset_dir_path']

class DataExporter:
    def __init__(self, config_json):
        self.config = DataExporterConfig(config_json)
        self.cached_schemas = {} # { 파일명 : 스키마 }
    
    def run(self):
        for schema_file_name in os.listdir(self.config.schema_dir_path):
            if schema_file_name.endswith('.table.json'):
                table_name = schema_file_name.split('.', 1)[0]
                #print(table_name)

                self.export_table(table_name)
    
    def export_table(self, table_name):
        schema = self.load_table_schema(table_name)

        type_info_root = self.create_type_info(schema)
        #print(str(type_info_root))

        data = self.read_table(type_info_root, table_name)
        #print(data)

        self.validate_table_data(data, schema)

        self.write_asset(data, table_name)

    def load_schema(self, schema_file_name, save_to_cache):
        if schema_file_name in self.cached_schemas:
            return self.cached_schemas[schema_file_name]

        schema_file_path = os.path.join(self.config.schema_dir_path, schema_file_name)

        schema = None
        with open(schema_file_path, 'r', encoding='UTF8') as fp:
            schema = json.load(fp)

        Draft7Validator.check_schema(schema)

        if save_to_cache:
            self.cached_schemas[schema_file_name] = schema

        return schema

    def load_table_schema(self, table_name):
        schema_file_name = table_name + '.table.json'
        return self.load_schema(schema_file_name, False)

    ##################################################
    # 타입 정보 생성
    ##################################################

    def create_type_info(self, schema):
        root = self.create_type_node('_root', schema)
        return root

    def create_type_node(self, node_name, node_schema):
        node_schema = self.resolve_schema_ref(node_schema)
        node = TypeNode()
        node.name = node_name

        node_type = node_schema['type']
        
        if node_type == 'array':
            node.is_array = True
            
            item_schema = self.resolve_schema_ref(node_schema['items'])
            item_type = item_schema['type']

            if item_type == 'array':
                raise Exception('Nested array schema is not supported!')
            elif item_type == 'object':
                for item_property_name, item_property_schema in item_schema['properties'].items():
                    node.add_member(self.create_type_node(item_property_name, item_property_schema))
            else:
                node.parse_type = self.determine_leaf_node_parse_type(item_type)
        
        elif node_type == 'object':
            for node_property_name, node_property_schema in node_schema['properties'].items():
                    node.add_member(self.create_type_node(node_property_name, node_property_schema))
        
        else:
            node.parse_type = self.determine_leaf_node_parse_type(node_type)

        return node

    def determine_leaf_node_parse_type(self, node_type):
        if node_type == 'integer':
            return TypeNodeParseType.INT
        elif node_type == 'number':
            return TypeNodeParseType.FLOAT
        elif node_type == 'boolean':
            return TypeNodeParseType.BOOL
        elif node_type == 'string':
            return TypeNodeParseType.STRING
        else:
            return TypeNodeParseType.NONE

    def resolve_schema_ref(self, schema):
        if '$ref' in schema:
            ref_uri = schema['$ref']    # enum.schema.json#/definitions/dayOfWeek
            ref_schema_file_name, ref_schema_path = ref_uri.split('#')  # 'enum.schema.json', '/definitions/dayOfWeek'
            
            ref_schema = self.load_schema(ref_schema_file_name, True)
            
            paths = ref_schema_path.split('/') # ['', 'definitions', 'dayOfWeek']
            ref_schema_node = ref_schema
            for i in range(1, len(paths)):
                ref_schema_node = ref_schema_node[paths[i]]
            
            return self.resolve_schema_ref(ref_schema_node)

        else:
            return schema

    ##################################################
    # 엑셀 테이블을 데이터로 변환
    ##################################################

    def read_table(self, type_info_root, table_name):
        table_file_name = table_name + '.xlsx'
        table_file_path = os.path.join(self.config.table_dir_path, table_file_name)

        wb = load_workbook(table_file_path, data_only=True)
        ws_data = wb['Data']
        type_max_depth = DataExporter.calc_max_depth(type_info_root, 0)
        header_info_root = DataExporter.create_data_header_info(ws_data, type_max_depth)
        #print(str(header_info_root))

        DataExporter.validate_header_info(type_info_root, header_info_root)

        data = DataExporter.load_datas(ws_data, type_info_root, header_info_root)
        #print(data)

        return data

    # 특정 노드 하위에서 지정한 이름 경로의 타입 정보 노드를 찾는다.
    # @param names 예) ['rewards', 'condition', 'firstClear']
    @staticmethod
    def find_node_py_path(node, node_path):
        if len(node_path) == 0:
            return node
            
        member = node.find_member(node_path[0])
        if member is not None:
            return DataExporter.find_node_py_path(member, node_path[1:])

        return None

    # 타입 정보 트리에서 가장 깊은 depth 값을 구한다.
    @staticmethod
    def calc_max_depth(node, cur_depth):
        if node is None:
            return cur_depth
        
        max_depth = cur_depth
        for member in node.members:
            member_depth = DataExporter.calc_max_depth(member, cur_depth + 1)
            
            if member_depth > max_depth:
                max_depth = member_depth
        
        return max_depth

    @staticmethod
    def calc_node_col_end(ws, row_idx, col_idx_start, col_idx_max):
        # 시작 컬럼은 항상 유효하다고 가정
        # start + 1 ~ max 순회
        # 값이 있는 다음 셀 찾아 그 전까지를 노드 마지막 col 로 구함
        for col_idx in range(col_idx_start + 1, col_idx_max + 1):
            if ws.cell(row=row_idx, column=col_idx).value is not None:
                return col_idx - 1
        
        return col_idx_max
    
    @staticmethod
    def calc_node_row_end(ws, col_idx, row_idx_start, row_idx_max):
        # 시작 행은 항상 유효하다고 가정
        # start + 1 ~ max 순회
        # 값이 있는 다음 셀 찾아 그 전까지를 노드 마지막 row 로 구함
        for row_idx in range(row_idx_start + 1, row_idx_max + 1):
            if ws.cell(row=row_idx, column=col_idx).value is not None:
                return row_idx - 1
        
        return row_idx_max
    
    @staticmethod
    def is_empty_row(ws, row_idx, col_idx_start, col_idx_max):
        for row in ws.iter_rows(min_row=row_idx, max_row=row_idx, min_col=col_idx_start, max_col=col_idx_max):
            for cell in row:
                if cell.value is not None:
                    return False
        
        return True

    @staticmethod
    def create_data_header_info(ws, max_depth):
        root = HeaderNode('_root', 1, ws.max_column)
        ancestors = [root]

        for col in ws.iter_cols(max_row=max_depth):
            for cell in col:
                if cell.value is not None:
                    parent = ancestors[-1]
                    parent_depth = len(ancestors) - 1
                    cell_depth = cell.row

                    node_name = cell.value
                    node_col_start = cell.col_idx

                    if cell_depth == max_depth:
                        # 자식 leaf 노드
                        node_col_end = DataExporter.calc_node_col_end(ws, cell.row, node_col_start, parent.col_end)
                        newNode = HeaderNode(node_name, node_col_start, node_col_end)
                        parent.add_member(newNode)
                    
                    elif cell_depth == parent_depth + 1:
                        # 자식 중간 노드
                        node_col_end = DataExporter.calc_node_col_end(ws, cell.row, node_col_start, parent.col_end)
                        newNode = HeaderNode(node_name, node_col_start, node_col_end)
                        parent.add_member(newNode)
                        ancestors.append(newNode)
                    
                    elif cell_depth == parent_depth:
                        # 부모와 같은 depth 의 중간 노드
                        del ancestors[-1]
                        newParent = ancestors[-1]
                        node_col_end = DataExporter.calc_node_col_end(ws, cell.row, node_col_start, newParent.col_end)
                        newNode = HeaderNode(node_name, node_col_start, node_col_end)
                        newParent.add_member(newNode)
                        ancestors.append(newNode)
                    
                    elif cell_depth < parent_depth:
                        # 부모를 거슬러 올라가는 depth 의 중간노드
                        up_depth = cell_depth - parent_depth - 1
                        del ancestors[up_depth:]
                        newParent = ancestors[-1]
                        node_col_end = DataExporter.calc_node_col_end(ws, cell.row, node_col_start, newParent.col_end)
                        newNode = HeaderNode(node_name, node_col_start, node_col_end)
                        newParent.add_member(newNode)
                        ancestors.append(newNode)

        return root
    
    @staticmethod
    def validate_header_info(type_root, header_root):
        # 모든 타입 노드가 헤더 노드에도 있어야 함
        # 헤더 노드에는 주석 컬럼 등도 있을 수 있으므로 반대는 체크하지 않음
        DataExporter.validate_type_node_exist_in_header(type_root, header_root, [])
    
    @staticmethod
    def validate_type_node_exist_in_header(type_node, header_root, parent_node_path):
        for member in type_node.members:
            node_path = parent_node_path + [member.name]
            
            if DataExporter.find_node_py_path(header_root, node_path) is None:
                raise Exception('Node {} is not exist in data header'.format(node_path))

            DataExporter.validate_type_node_exist_in_header(member, header_root, node_path)

    @staticmethod
    def parse_cell(parse_type, cell):
        try:
            # 타입 무관하게 셀이 비어있으면 None 취급
            if cell.value is None:
                return None

            if parse_type == TypeNodeParseType.INT:
                return int(cell.value)
            elif parse_type == TypeNodeParseType.FLOAT:
                return float(cell.value)
            elif parse_type == TypeNodeParseType.BOOL:
                return bool(cell.value)
            elif parse_type == TypeNodeParseType.STRING:
                return str(cell.value) if cell.value is not None else ''
            else:
                raise Exception('Invalid type name to parse. parse_type:{}, coordinate:{}'.format(parse_type, cell.coordinate))
        except Exception as e:
            print('Exception on parse_cell(). coordinate:{}'.format(cell.coordinate))
            raise e

    @staticmethod
    def load_datas(ws, type_info_root, header_info_root):
        header_row_count = DataExporter.calc_max_depth(type_info_root, 0)
        data_row_start = header_row_count + 1
        data_row_max = max(ws.max_row, data_row_start)

        return DataExporter.load_node_array(ws, type_info_root, header_info_root, data_row_start, data_row_max)

    @staticmethod
    def load_node_object(ws, type_node, header_node, row_start, row_max):
        data = {}
        row_end = row_max
        
        for member_header_node in header_node.members:
            member_name = member_header_node.name
            member_type_node = type_node.find_member(member_header_node.name)

            # 주석 컬럼은 타입 정보에 없음. 오류 아님
            if member_type_node is None:
                continue
            
            # 배열은 여러 행에 걸쳐 있을 수 있으므로
            # 배열이 아닌 타입을 기준으로 노드 마지막 행을 갱신한다.
            if not member_type_node.is_array:
                row_end = DataExporter.calc_node_row_end(ws, member_header_node.col_start, row_start, row_end)

            if member_type_node.is_leaf():
                if not member_type_node.is_array:
                    # 단일값
                    data_value = DataExporter.load_leaf_value(ws, member_type_node, member_header_node, row_start)
                    if not data_value is None:
                        data[member_name] = data_value
                else:
                    # 단일값 배열
                    data_values = DataExporter.load_leaf_array(ws, member_type_node, member_header_node, row_start, row_end)
                    if not data_values is None:
                        if not member_name in data:
                            data[member_name] = []
                        data[member_name].extend(data_values)
            else:
                if not member_type_node.is_array:
                    # 구조체
                    data_value = DataExporter.load_node_object(ws, member_type_node, member_header_node, row_start, row_end).data
                    if not data_value is None:
                        data[member_name] = data_value
                else:
                    # 구조체 배열
                    data_values = DataExporter.load_node_array(ws, member_type_node, member_header_node, row_start, row_end)
                    if not data_values is None:
                        if not member_name in data:
                            data[member_name] = []
                        data[member_name].extend(data_values)

        if len(data) == 0:
            data = None
        return LoadedNodeObjectInfo(data, row_end)

    @staticmethod
    def load_node_array(ws, type_node, header_node, row_start, row_max):
        datas = []
        row_idx = row_start

        while row_idx <= row_max:
            # 빈 줄을 만나면 다음 줄로 넘어감
            # 데이터 구분 위해 의도적으로 빈 줄 남기는 경우 위함
            if DataExporter.is_empty_row(ws, row_idx, header_node.col_start, header_node.col_end):
                row_idx += 1
                continue
            
            loaded_info = DataExporter.load_node_object(ws, type_node, header_node, row_idx, row_max)
            if not loaded_info.data is None:
                datas.append(loaded_info.data)
            row_idx = loaded_info.row_end + 1

        if len(datas) > 0:
            return datas
        else:
            return None

    # 리프노드 단일값 구하기
    @staticmethod
    def load_leaf_value(ws, type_node, header_node, row_start):
        cell = ws.cell(column=header_node.col_start, row=row_start)
        return DataExporter.parse_cell(type_node.parse_type, cell)

    # 리프노드 배열 구하기
    @staticmethod
    def load_leaf_array(ws, type_node, header_node, row_start, row_max):
        datas = []
        
        for col in ws.iter_cols(min_col=header_node.col_start, max_col=header_node.col_start, min_row=row_start, max_row=row_max):
            for cell in col:
                # 빈 줄을 만나면 다음 줄로 넘어감
                # load_node_array() 와 동작 맞추기 위함
                if cell.value is None:
                    continue
                
                data_value = DataExporter.parse_cell(type_node.parse_type, cell)
                if not data_value is None:
                    datas.append(data_value)
        
        if len(datas) > 0:
            return datas
        else:
            return None


    ##################################################
    # 테이블 데이터 유효성 검사
    ##################################################

    def validate_table_data(self, data, schema):
        schema_dir_abs_path = os.path.abspath(self.config.schema_dir_path)
        resolver = RefResolver('file:{}/'.format(schema_dir_abs_path), None)

        validator = Draft7Validator(schema, resolver=resolver)
        validator.validate(data)


    ##################################################
    # 데이터 어셋 파일 쓰기
    ##################################################

    def write_asset(self, data, table_name):
        asset_file_name = table_name + '.json'
        asset_file_path = os.path.join(self.config.asset_dir_path, asset_file_name)

        old_dump = None

        try:
            with open(asset_file_path, 'r', encoding='utf8') as datafile:
                old_dump = datafile.read()
        except FileNotFoundError:
            pass

        new_dump = json.dumps(data, indent=4, ensure_ascii=False)

        write_status = ''
        
        if old_dump != new_dump:
            with open(asset_file_path, 'w', encoding='utf8') as datafile:
                datafile.write(new_dump)
            write_status = 'Written'
        else:
            write_status = 'Skipped'

        print('{}: {}'.format(asset_file_path, write_status))